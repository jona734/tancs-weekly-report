import os
import requests
import json
from datetime import datetime, date
from base64 import b64encode
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 환경변수 ──────────────────────────────────────────────
BASE_URL   = os.environ["JIRA_BASE_URL"].rstrip("/")
EMAIL      = os.environ["JIRA_EMAIL"]
API_TOKEN  = os.environ["JIRA_API_TOKEN"]

AUTH = b64encode(f"{EMAIL}:{API_TOKEN}".encode()).decode()
HEADERS = {
    "Authorization": f"Basic {AUTH}",
    "Accept": "application/json",
    "Content-Type": "application/json",
}

# ── 대상 프로젝트 (TANCS7 제외) ───────────────────────────
PROJECTS = ["TANCS", "TANCS1", "TANCS2", "TANCS3", "TANCS4", "TANCS5", "TANCS6"]

# ── 제외 issue_type ──────────────────────────────────────
EXCLUDE_TYPES = {"Sub-task", "IAR", "FAR", "CBC", "IAR/FAR/CBC"}

# ── JQL ─────────────────────────────────────────────────
START_DATE = "2026-01-01"
TODAY = date.today().isoformat()
PROJECT_JQL = " OR ".join([f'project = "{p}"' for p in PROJECTS])
JQL = (
    f'({PROJECT_JQL}) '
    f'AND created >= "{START_DATE}" '
    f'AND created <= "{TODAY}" '
    f'AND issuetype not in ("Sub-task","IAR","FAR","CBC") '
    f'ORDER BY created DESC'
)

# ── 가져올 필드 ──────────────────────────────────────────
FIELDS = [
    "summary", "status", "assignee", "reporter", "updated", "created",
    "duedate", "labels", "issuelinks", "security", "resolution", "issuetype",
    "priority",
    # TANCS 커스텀 필드 (실제 field ID는 아래에서 자동 탐색)
    "customfield_titan_project_name",
    "customfield_chip",
    "customfield_cust_application",
    "customfield_sdk_version_titan",
    "customfield_sub_device_multi",
    "customfield_ref_hw_version",
    "customfield_os",
    "customfield_self_resolution",
    "customfield_start_date",
    "customfield_fae_person",
    "customfield_cause_customer",
    "customfield_hw_issue_pattern",
    "customfield_sw_issue_pattern",
]

# ── 커스텀 필드 ID 매핑 (JIRA 인스턴스마다 다를 수 있음) ──
# fetch_field_map() 에서 실제 ID를 조회해서 채움
FIELD_MAP = {}

def fetch_field_map():
    """JIRA 필드 목록 조회 → name → id 매핑"""
    url = f"{BASE_URL}/rest/api/3/field"
    r = requests.get(url, headers=HEADERS, timeout=30)
    r.raise_for_status()
    for f in r.json():
        FIELD_MAP[f.get("name", "").lower().replace(" ", "_")] = f["id"]
        FIELD_MAP[f["id"]] = f.get("name", f["id"])

def resolve_custom_field(name_key: str) -> str:
    """이름 키로 실제 field ID 반환; 없으면 원본 반환"""
    return FIELD_MAP.get(name_key, name_key)

def fetch_issues() -> list:
    """페이지네이션으로 전체 이슈 수집"""
    issues = []
    start = 0
    page_size = 100

    # 커스텀 필드 ID 목록 구성
    field_ids = [
        "summary", "status", "assignee", "reporter", "updated", "created",
        "duedate", "labels", "issuelinks", "security", "resolution", "issuetype", "priority",
    ]
    # 커스텀 필드명 → ID 변환
    custom_names = {
        "titan_project_name": None,
        "chip": None,
        "cust_application": None,
        "sdk_version_titan": None,
        "sub_device_multi": None,
        "ref_hw_version": None,
        "os": None,
        "self_resolution": None,
        "start_date": None,
        "fae_person": None,
        "cause_customer": None,
        "hw_issue_pattern": None,
        "sw_issue_pattern": None,
    }
    for key in custom_names:
        cid = FIELD_MAP.get(key) or FIELD_MAP.get(key.replace("_", " "))
        if cid:
            custom_names[key] = cid
            field_ids.append(cid)

    print(f"[INFO] JQL: {JQL}")

    while True:
        url = f"{BASE_URL}/rest/api/3/search"
        payload = {
            "jql": JQL,
            "startAt": start,
            "maxResults": page_size,
            "fields": field_ids,
        }
        r = requests.post(url, headers=HEADERS, json=payload, timeout=60)
        r.raise_for_status()
        data = r.json()
        batch = data.get("issues", [])
        issues.extend(batch)
        total = data.get("total", 0)
        start += len(batch)
        print(f"[INFO] Fetched {start}/{total}")
        if start >= total or not batch:
            break

    return issues, custom_names

def safe_get(obj, *keys, default=""):
    """중첩 dict 안전 탐색"""
    for k in keys:
        if not isinstance(obj, dict):
            return default
        obj = obj.get(k, {})
    return obj if obj not in (None, {}, []) else default

def parse_sw_pattern(value: str):
    """
    sw_issue_pattern 계층 분리
    예: "SW_Bug > Driver Issue" → ("SW_Bug", "Driver Issue")
    구분자 없으면 → (value, "")
    """
    if not value:
        return "", ""
    for sep in [" > ", ">", " / ", "/"]:
        if sep in value:
            parts = value.split(sep, 1)
            return parts[0].strip(), parts[1].strip()
    return value.strip(), ""

def extract_row(issue: dict, custom_names: dict) -> dict:
    f = issue.get("fields", {})

    # 기본 필드
    row = {
        "id":               issue.get("id", ""),
        "create_date":      (f.get("created") or "")[:10],
        "issue_key":        issue.get("key", ""),
        "summary":          f.get("summary", ""),
        "status":           safe_get(f, "status", "name"),
        "assignee":         safe_get(f, "assignee", "displayName"),
        "reporter":         safe_get(f, "reporter", "displayName"),
        "updated":          (f.get("updated") or "")[:10],
        "created":          (f.get("created") or "")[:10],
        "duedate":          f.get("duedate", ""),
        "labels":           ", ".join(f.get("labels") or []),
        "issue_links":      "; ".join([
                                safe_get(lnk, "outwardIssue", "key") or safe_get(lnk, "inwardIssue", "key")
                                for lnk in (f.get("issuelinks") or [])
                            ]),
        "security_level":   safe_get(f, "security", "name"),
        "resolution":       safe_get(f, "resolution", "name"),
        "issue_type":       safe_get(f, "issuetype", "name"),
        "priority":         safe_get(f, "priority", "name"),
    }

    # 커스텀 필드 매핑
    def cf(key):
        fid = custom_names.get(key)
        if not fid:
            return ""
        val = f.get(fid)
        if val is None:
            return ""
        if isinstance(val, dict):
            return val.get("value") or val.get("name") or val.get("displayName") or str(val)
        if isinstance(val, list):
            parts = []
            for v in val:
                if isinstance(v, dict):
                    parts.append(v.get("value") or v.get("name") or str(v))
                else:
                    parts.append(str(v))
            return ", ".join(parts)
        return str(val)

    row["titan_project_name"] = cf("titan_project_name")
    row["chip"]               = cf("chip")
    row["cust_application"]   = cf("cust_application")
    row["sdk_version_titan"]  = cf("sdk_version_titan")
    row["sub_device_multi"]   = cf("sub_device_multi")
    row["ref_hw_version"]     = cf("ref_hw_version")
    row["os"]                 = cf("os")
    row["self_resolution"]    = cf("self_resolution")
    row["start_date"]         = cf("start_date")
    row["fae_person"]         = cf("fae_person")
    row["cause_customer"]     = cf("cause_customer")
    row["hw_issue_pattern"]   = cf("hw_issue_pattern")

    sw_raw = cf("sw_issue_pattern")
    l1, l2 = parse_sw_pattern(sw_raw)
    row["sw_issue_pattern_L1"] = l1
    row["sw_issue_pattern_L2"] = l2

    return row

def apply_filter(row: dict) -> bool:
    """True = 유지, False = 제외"""
    if not row.get("titan_project_name", "").strip():
        return False
    if row.get("issue_type", "") in EXCLUDE_TYPES:
        return False
    return True

# ── 컬럼 정의 ────────────────────────────────────────────
COLUMNS = [
    ("id",                   "ID",                    12),
    ("create_date",          "Create Date",           14),
    ("issue_key",            "Issue Key",             14),
    ("summary",              "Summary",               50),
    ("status",               "Status",                14),
    ("assignee",             "Assignee",              18),
    ("reporter",             "Reporter",              18),
    ("updated",              "Updated",               14),
    ("created",              "Created",               14),
    ("duedate",              "Due Date",              14),
    ("labels",               "Labels",                20),
    ("issue_links",          "Issue Links",           20),
    ("security_level",       "Security Level",        16),
    ("resolution",           "Resolution",            16),
    ("issue_type",           "Issue Type",            16),
    ("priority",             "Priority",              12),
    ("titan_project_name",   "Titan Project Name",    20),
    ("chip",                 "Chip",                  14),
    ("cust_application",     "Cust Application",      20),
    ("sdk_version_titan",    "SDK Version (Titan)",   20),
    ("sub_device_multi",     "Sub Device Multi",      18),
    ("ref_hw_version",       "Ref HW Version",        18),
    ("os",                   "OS",                    14),
    ("self_resolution",      "Self Resolution",       18),
    ("start_date",           "Start Date",            14),
    ("fae_person",           "FAE Person",            16),
    ("cause_customer",       "Cause (Customer)",      20),
    ("hw_issue_pattern",     "HW Issue Pattern",      20),
    ("sw_issue_pattern_L1",  "SW Issue Pattern (L1)", 22),
    ("sw_issue_pattern_L2",  "SW Issue Pattern (L2)", 22),
]

def build_excel(rows: list, output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TANCS Weekly"

    # 스타일 정의
    header_font    = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill    = PatternFill("solid", start_color="1F4E79")  # 진청
    sub_fill       = PatternFill("solid", start_color="2E75B6")  # 중간 청 (sw 계층 컬럼)
    center_align   = Alignment(horizontal="center", vertical="center", wrap_text=False)
    wrap_align     = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    left_align     = Alignment(horizontal="left",   vertical="center")
    thin           = Side(style="thin", color="D0D0D0")
    border         = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 헤더 행
    ws.row_dimensions[1].height = 28
    for col_idx, (key, label, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font      = header_font
        cell.fill      = sub_fill if "SW Issue Pattern" in label else header_fill
        cell.alignment = center_align
        cell.border    = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 데이터 행
    alt_fill = PatternFill("solid", start_color="EBF3FB")  # 연청 줄무늬
    for row_idx, row in enumerate(rows, start=2):
        ws.row_dimensions[row_idx].height = 18
        is_alt = (row_idx % 2 == 0)
        for col_idx, (key, _, _) in enumerate(COLUMNS, start=1):
            val = row.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = Font(name="Arial", size=9)
            cell.alignment = wrap_align if key == "summary" else left_align
            cell.border    = border
            if is_alt:
                cell.fill = alt_fill

    # 헤더 고정 및 자동 필터
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    # 요약 시트
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "추출 기준"
    ws2["B1"] = f"{START_DATE} ~ {TODAY}"
    ws2["A2"] = "총 이슈 수"
    ws2["B2"] = f'=COUNTA(\'TANCS Weekly\'!A2:A10000)-SUMPRODUCT((\'TANCS Weekly\'!A2:A10000="")*1)'
    ws2["A3"] = "생성일시"
    ws2["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws2["A4"] = "대상 프로젝트"
    ws2["B4"] = ", ".join(PROJECTS)

    for cell in [ws2["A1"], ws2["A2"], ws2["A3"], ws2["A4"]]:
        cell.font = Font(name="Arial", bold=True, size=10)

    wb.save(output_path)
    print(f"[INFO] 저장 완료: {output_path} ({len(rows)}건)")

def main():
    print("[START] TANCS Weekly Report 생성 시작")
    fetch_field_map()
    print(f"[INFO] 필드 맵 로드 완료 ({len(FIELD_MAP)} 항목)")

    issues, custom_names = fetch_issues()
    print(f"[INFO] 원본 이슈 수: {len(issues)}")

    rows = []
    for issue in issues:
        row = extract_row(issue, custom_names)
        if apply_filter(row):
            rows.append(row)

    print(f"[INFO] 필터 후 이슈 수: {len(rows)}")

    today_str = date.today().strftime("%Y%m%d")
    output_path = f"TANCS_Weekly_{today_str}.xlsx"
    build_excel(rows, output_path)
    print("[DONE] 완료")

if __name__ == "__main__":
    main()
